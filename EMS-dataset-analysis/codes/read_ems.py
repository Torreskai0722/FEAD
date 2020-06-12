# -*- coding: utf-8 -*-

import pandas as pd
import pymysql
import json
import matplotlib.pyplot as plt
import numpy as np
from sklearn import datasets, linear_model
from sklearn.metrics import mean_squared_error, r2_score
import openpyxl
import csv 
 
# def write_excel_xlsx(path, sheet_name, value):
#     index = len(value)
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = sheet_name
#     for i in range(0, index):
#         for j in range(0, len(value[i])):
#             sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
#     workbook.save(path)
 
 
# def read_excel_xlsx(path, sheet_name):
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook[sheet_name]
#     for row in sheet.rows:
#         for cell in row:
#             print(cell.value, "\t", end="")
#         print()

def write_csv():
    path  = "aa.csv"
    with open(path,'a+') as f:
        csv_write = csv.writer(f)
        data_row = ["1","2"]
        csv_write.writerow(data_row)
 
 
book_name_vvid = '257C0D741E2CDDAFDA1A297FC5AC9964.xlsx'
 
sheet_name_date = '20191201'
sheet_name_label = '20191201-label' 
 
# write_excel_xlsx(book_name_vvid, sheet_name_date, datavalue)
# read_excel_xlsx(book_name_vvid, sheet_name_date)

## 加上字符集参数，防止中文乱码
dbconn=pymysql.connect(
  host="localhost",
  database="fuel_efficiency",
  user="root",
  password="123456",
  port=3306,
  charset='utf8'
 )

#sql语句
sqlcmd="SELECT ems_value,lat,lng FROM emsdata where (truckid = '257C0D741E2CDDAFDA1A297FC5AC9964') AND (tdate = '20191201')"

#利用pandas 模块导入mysql数据
a=pd.read_sql(sqlcmd,dbconn)
#取前5行数据

X = []
y = []

# b = a
b = a.head()
for i in range(len(b.index)):
	print(i)
	d = json.loads(b['ems_value'][i])
	data_ems = [float(d.get('x7000', '0')),float(d.get('x7002', '0')),float(d.get('x7003', '0')),float(d.get('x7004', '0')), 
		float(d.get('x7005', '0')),float(d.get('x7006', '0')),float(d.get('x7007', '0')),float(d.get('x006c', '0')),float(d.get('x7035', '0')),float(d.get('x7091', '0')),float(b['lat'][i]),float(b['lng'][i])]
	data_label = [float(d.get('x7001', '0'))]
	print(type(d.get('x7001', '0')))
	X.append(data_ems)
	y.append(data_label)
	# write_excel_xlsx(book_name_vvid, sheet_name_date, data_ems)
	# write_excel_xlsx(book_name_vvid, sheet_name_label, data_label)
	# print(d.get('x7000', '0'),d.get('x7001', '0'),d.get('x7002', '0'),d.get('x7003', '0'),d.get('x7004', '0'), 
	# 	d.get('x7005', '0'),d.get('x7006', '0'),d.get('x7007', '0'),d.get('x006c', '0'),d.get('x7035', '0'),d.get('x7091', '0'))
	# print(b['lat'][i])
	# print(b['lng'][i])

# xls = pd.ExcelFile(book_name_vvid)
# df1 = pd.read_excel(xls, sheet_name_date)
# X = np.array(df1)
# df2 = pd.read_excel(xls, sheet_name_label)
# y = np.array(df2)
print(len(b.index))
print(X)
print(y)

# np_d = b.values
# test = b.to_string()
# for r in b:
# 	c = r.to_json(orient='table')
# 	print(c)

# print(b)
# print(type(b))
# print(b.shape)
# print(b.info())
# print(b.dtypes)
# print(b['ems_value'][0])
# print(b['ems_value'][1])
# print(type(b['ems_value'][1]))
# print(b['lat'])
# print(b['lng'])

# d = json.loads(b['ems_value'][1])
# print(d['x7000'],d['x7001'],d['x7002'],d['x7003'],d['x7004'],d['x7005'],d['x7006'],d['x7007'],d['x006C'],d['x7035'],d['x7091'],d['x70EB'])
# print(np_d)
# print(a[''])
# b=a.head(50)
# print(b)

# 读取csv数据
# pd.read_csv()

# 读取excel数据
# xls = pd.ExcelFile(book_name_vvid)
# df1 = pd.read_excel(xls, sheet_name_date)
# X = np.array(df1)
# df2 = pd.read_excel(xls, sheet_name_label)
# y = np.array(df2)
# tt = pd.read_excel('257C0D741E2CDDAFDA1A297FC5AC9964.xlsx')
# WS_np = np.array(tt)
# print(tt)

# 读取txt数据
#pd.read_table()

# VVID								License Plate	Days
# 3CCC005122531737E69EA3BA21324ECD 	粤ABP691			1:4, 6:31
# A0A4A31F4C3509655240D8D7DB9CD389	闽K59769			1:4, 6:31
# 83E94E04A08895767DFE0D80A21A07D3 	闽K59938			1:4, 6:31
# B45F36E3944670E22C7EC735D833F709	粤ADW293			27:31
# 7571522FF0EBA036818CEACBA52D3B60 	粤ADS670			28:31
# 096B3BBA5216C10C7EDF72FD803ACFD7 	粤ADP980			3:4, 9:31
# 3EF3F915B5831AE8667B6FC54FBA89B7 	闽K55572			1:4, 6:31
# CABB5C23A2B5E1541DB6E75FD1D61E01 	粤ABW222			1:4, 6:31
# 257C0D741E2CDDAFDA1A297FC5AC9964	闽K59936			1:4, 6:31